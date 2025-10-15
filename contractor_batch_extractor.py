#!/usr/bin/env python3
"""
Batch Contractor Extractor
===========================
Extracts MULTIPLE contractors from a single Excel file

Assumes Excel structure:
- Header row with column names
- One contractor per row
- Columns: Name, Email, Phone, Address, Postcode, Services, etc.
"""

import re
import json
import openpyxl
from pathlib import Path
from typing import Dict, List
from datetime import datetime


class BatchContractorExtractor:
    """Extract multiple contractors from Excel file"""
    
    def __init__(self):
        self.contractors = []
        
        # Common column name variations (exact matches for common Excel columns)
        self.column_mappings = {
            'name': ['^name$', 'contractor name', 'company name', 'supplier name'],
            'email': ['^email', 'e-mail', 'email address', 'contact email'],
            'phone': ['telephone', '^mobile$', 'phone', 'tel', 'contact'],
            'address': ['^address$', 'street address', 'full address', 'location'],
            'postcode': ['^postcode$', 'post code', 'zip', 'postal code'],
            'services': ['services', 'service', 'services provided', 'trade'],
            'bank_account': ['payments bank account', 'bank account', 'account number', 'account name'],
            'sort_code': ['payments bank sort code', 'bank sort code', 'sort code', 'sortcode'],
            'pli_expiry': ['pli expiry', 'insurance expiry', 'pli date', 'expiry'],
            'accounts': ['accounts', 'audited accounts', 'financial statements'],
            'incorporation': ['co registration', 'registration', 'incorporation', 'companies house']
        }
    
    def extract_from_excel(self, file_path: str) -> List[Dict]:
        """
        Extract all contractors from Excel file
        
        Returns:
            List of contractor dictionaries
        """
        file = Path(file_path)
        
        print(f"\n📊 Extracting multiple contractors from: {file.name}")
        
        try:
            # Open workbook
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
            
            print(f"   Worksheet: {ws.title}")
            print(f"   Dimensions: {ws.max_row} rows × {ws.max_column} columns")
            
            # Find header row and map columns
            header_row = None
            column_map = {}
            
            for row_idx in range(1, min(10, ws.max_row + 1)):
                row_values = [cell.value for cell in ws[row_idx]]
                row_text = ' '.join([str(v).lower() for v in row_values if v])
                
                # Check if this looks like a header row
                if any(keyword in row_text for keyword in ['name', 'email', 'contractor', 'company']):
                    header_row = row_idx
                    
                    # Map columns - check all columns for all fields
                    for col_idx, cell in enumerate(ws[row_idx], 1):
                        if cell.value:
                            col_name = str(cell.value).strip().lower()
                            
                            # Match to our field types (only if not already mapped)
                            for field, variations in self.column_mappings.items():
                                if field in column_map:
                                    continue  # Skip if already mapped
                                
                                for var in variations:
                                    matched = False
                                    # Check for regex pattern (starts with ^)
                                    if var.startswith('^'):
                                        if re.match(var, col_name):
                                            column_map[field] = col_idx
                                            print(f"      Mapped '{col_name}' → {field} (col {col_idx})")
                                            matched = True
                                    # Simple substring match
                                    elif var in col_name:
                                        column_map[field] = col_idx
                                        print(f"      Mapped '{col_name}' → {field} (col {col_idx})")
                                        matched = True
                                    
                                    if matched:
                                        break
                    
                    print(f"\n   ✅ Header row found at row {header_row}")
                    print(f"   📋 Mapped columns: {list(column_map.keys())}")
                    break
            
            if not header_row:
                print(f"   ⚠️  No header row found, trying row-by-row extraction...")
                return self._extract_without_headers(ws)
            
            # Extract contractors from data rows
            contractors_found = 0
            empty_rows = 0
            
            name_col = column_map.get('name')
            if not name_col:
                print(f"   ⚠️  No name column mapped!")
                return []
            
            print(f"\n   Scanning rows {header_row + 1} to {min(header_row + 100, ws.max_row + 1)}...")
            print(f"   Name column: {name_col}")
            
            for row_idx in range(header_row + 1, min(header_row + 100, ws.max_row + 1)):  # First 100 rows
                
                # Get cell value (column index is 1-based)
                name_cell = ws.cell(row=row_idx, column=name_col)
                name_value = name_cell.value
                
                if row_idx <= header_row + 5:
                    print(f"   Debug row {row_idx}: name_col={name_col}, value='{name_value}'")
                
                if not name_value or str(name_value).strip() == '':
                    empty_rows += 1
                    if empty_rows > 20:
                        break  # Stop after 20 consecutive empty rows
                    continue
                
                empty_rows = 0  # Reset counter
                contractors_found += 1
                
                if contractors_found <= 5:
                    print(f"   ✓ Row {row_idx}: {name_value}")
                
                # Extract contractor data using ws.cell()
                def get_cell_value(field_key):
                    if field_key in column_map:
                        return ws.cell(row=row_idx, column=column_map[field_key]).value
                    return None
                
                contractor = {
                    'contractor_name': self._clean_value(name_value),
                    'email': self._clean_value(get_cell_value('email')),
                    'telephone': self._clean_value(get_cell_value('phone')),
                    'address': self._clean_value(get_cell_value('address')),
                    'postcode': self._clean_value(get_cell_value('postcode')),
                    'services_provided': self._parse_services(get_cell_value('services')),
                    'bank_account_name': self._clean_value(get_cell_value('bank_account')),
                    'bank_sort_code': self._clean_value(get_cell_value('sort_code')),
                    'pli_expiry_date': self._parse_date(get_cell_value('pli_expiry')),
                    'has_audited_accounts': bool(get_cell_value('accounts')),
                    'has_certificate_of_incorporation': bool(get_cell_value('incorporation')),
                    'extraction_method': 'excel_batch',
                    'extraction_confidence': 0.8,
                    'row_number': row_idx
                }
                
                # Only add if has minimum data
                if contractor['contractor_name']:
                    self.contractors.append(contractor)
            
            wb.close()
            
            print(f"\n   ✅ Extracted {len(self.contractors)} contractors!")
            
            # Show summary
            for i, c in enumerate(self.contractors[:5], 1):
                print(f"      {i}. {c['contractor_name']}")
            
            if len(self.contractors) > 5:
                print(f"      ... and {len(self.contractors) - 5} more")
            
            return self.contractors
            
        except Exception as e:
            print(f"\n   ❌ Error: {e}")
            import traceback
            traceback.print_exc()
            return []
    
    def _extract_without_headers(self, ws) -> List[Dict]:
        """Extract using pattern matching when no clear headers"""
        print(f"   Using pattern-based extraction...")
        
        for row_idx in range(1, ws.max_row + 1):
            row_text = []
            for cell in ws[row_idx]:
                if cell.value:
                    row_text.append(str(cell.value))
            
            text = ' '.join(row_text)
            
            # Look for company names (Limited, Ltd, etc.)
            company_pattern = r'([A-Z][A-Za-z\s&]+(?:Limited|Ltd|LLP|plc))'
            match = re.search(company_pattern, text)
            
            if match:
                contractor = {
                    'contractor_name': match.group(1).strip(),
                    'email': None,
                    'telephone': None,
                    'extraction_method': 'pattern_match',
                    'row_number': row_idx
                }
                
                # Extract other fields from same row
                email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
                email_match = re.search(email_pattern, text)
                if email_match:
                    contractor['email'] = email_match.group(0)
                
                self.contractors.append(contractor)
        
        return self.contractors
    
    def _clean_value(self, value):
        """Clean cell value"""
        if value is None:
            return None
        return str(value).strip() if str(value).strip() else None
    
    def _parse_services(self, value):
        """Parse services from cell (may be comma-separated)"""
        if not value:
            return []
        
        text = str(value)
        # Split by comma, semicolon, or newline
        services = re.split(r'[,;\n]', text)
        return [s.strip().title() for s in services if s.strip()]
    
    def _parse_date(self, value):
        """Parse date from cell"""
        if not value:
            return None
        
        # If already a date object
        if hasattr(value, 'strftime'):
            return value.strftime('%Y-%m-%d')
        
        # If string, try to parse
        date_str = str(value).strip()
        for fmt in ['%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d/%m/%y']:
            try:
                dt = datetime.strptime(date_str, fmt)
                return dt.strftime('%Y-%m-%d')
            except:
                continue
        
        return date_str


def main():
    """CLI for batch extraction"""
    import sys
    from contractor_sql_generator import SupplierSQLGenerator
    
    if len(sys.argv) < 2:
        print("Usage: python3 contractor_batch_extractor.py <excel_file.xlsx>")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    
    # Extract all contractors
    extractor = BatchContractorExtractor()
    contractors = extractor.extract_from_excel(excel_file)
    
    if not contractors:
        print("\n❌ No contractors found")
        sys.exit(1)
    
    print("\n" + "="*80)
    print(f"📊 EXTRACTED {len(contractors)} CONTRACTORS")
    print("="*80)
    
    # Generate SQL for each
    output_file = Path('output') / f"all_suppliers_{datetime.now().strftime('%Y%m%d_%H%M%S')}.sql"
    output_file.parent.mkdir(exist_ok=True)
    
    sql_parts = ["BEGIN;\n"]
    
    for i, contractor in enumerate(contractors, 1):
        print(f"\n{i}. {contractor['contractor_name']}")
        print(f"   Email: {contractor.get('email') or 'N/A'}")
        print(f"   Phone: {contractor.get('telephone') or 'N/A'}")
        print(f"   Postcode: {contractor.get('postcode') or 'N/A'}")
        
        # Generate SQL
        generator = SupplierSQLGenerator()
        sql = generator.generate_sql(contractor)
        
        # Remove BEGIN/COMMIT from individual SQL
        sql_clean = sql.replace('BEGIN;', '').replace('COMMIT;', '').strip()
        sql_parts.append(f"\n-- Contractor {i}: {contractor['contractor_name']}\n{sql_clean}")
    
    sql_parts.append("\nCOMMIT;")
    
    # Save combined SQL
    with open(output_file, 'w') as f:
        f.write('\n'.join(sql_parts))
    
    print(f"\n" + "="*80)
    print(f"✅ SQL GENERATED FOR ALL {len(contractors)} CONTRACTORS")
    print("="*80)
    print(f"\n📄 File: {output_file}")
    print(f"\n🚀 To apply to database:")
    print(f"   python3 apply_with_new_credentials.py {output_file}")
    print()


if __name__ == '__main__':
    main()

