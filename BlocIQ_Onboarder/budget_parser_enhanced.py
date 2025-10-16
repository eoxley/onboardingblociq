"""
Enhanced Budget Parser
Extracts actual budget amounts from Excel budget files
"""

import openpyxl
import re
from typing import Dict, List, Any, Optional
from datetime import datetime


class BudgetParser:
    """Parse budget Excel files to extract line item amounts"""
    
    def __init__(self):
        self.budget_keywords = [
            'budget', 'expenditure', 'income', 'service charge',
            'reserve fund', 'sinking fund', 'total'
        ]
        
    def parse_budget_file(self, file_path: str) -> Dict[str, Any]:
        """
        Parse budget Excel file and extract amounts
        
        Returns:
            {
                'year': '2024-2025',
                'total_service_charge': 123456.78,
                'total_reserve_fund': 12345.67,
                'line_items': [
                    {'description': 'Management Fee', 'amount': 22515.0, 'category': 'admin'},
                    ...
                ]
            }
        """
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            
            budget_data = {
                'year': self._extract_year_from_filename(file_path),
                'total_service_charge': 0,
                'total_reserve_fund': 0,
                'line_items': []
            }
            
            # Find the budget table structure
            header_row = self._find_header_row(ws)
            if not header_row:
                return budget_data
            
            # Extract column positions
            columns = self._identify_columns(ws, header_row)
            
            # Extract line items
            line_items = self._extract_line_items(ws, header_row + 1, columns)
            budget_data['line_items'] = line_items
            
            # Calculate totals
            budget_data['total_service_charge'] = sum(
                item['amount'] for item in line_items 
                if item.get('category') != 'reserve_fund'
            )
            budget_data['total_reserve_fund'] = sum(
                item['amount'] for item in line_items 
                if item.get('category') == 'reserve_fund'
            )
            
            return budget_data
            
        except Exception as e:
            print(f"  ⚠️  Error parsing budget {file_path}: {e}")
            return {'line_items': []}
    
    def _extract_year_from_filename(self, filename: str) -> Optional[str]:
        """Extract year from filename like 'Budget YE24' or 'Budget 2024-2025'"""
        filename_lower = filename.lower()
        
        # Try patterns like "YE24", "YE2024"
        match = re.search(r'ye\s*(\d{2,4})', filename_lower)
        if match:
            year = match.group(1)
            if len(year) == 2:
                year = '20' + year
            return f"{year}-{int(year)+1}"
        
        # Try patterns like "2024-2025" or "2024"
        match = re.search(r'(\d{4})\s*-?\s*(\d{4})?', filename_lower)
        if match:
            start_year = match.group(1)
            end_year = match.group(2) or str(int(start_year) + 1)
            return f"{start_year}-{end_year}"
        
        return None
    
    def _find_header_row(self, ws) -> Optional[int]:
        """Find the row containing budget headers"""
        for row_idx in range(1, min(50, ws.max_row + 1)):
            row_text = ' '.join(
                str(ws.cell(row=row_idx, column=col).value or '').lower() 
                for col in range(1, min(20, ws.max_column + 1))
            )
            
            # Look for typical budget headers
            if any(keyword in row_text for keyword in ['description', 'item', 'category', 'expenditure']):
                if any(keyword in row_text for keyword in ['amount', 'total', '£', 'cost', 'budget']):
                    return row_idx
        
        return None
    
    def _identify_columns(self, ws, header_row: int) -> Dict[str, int]:
        """Identify which columns contain description and amount"""
        columns = {}
        
        for col in range(1, min(20, ws.max_column + 1)):
            header = str(ws.cell(row=header_row, column=col).value or '').lower().strip()
            
            if any(keyword in header for keyword in ['description', 'item', 'expense', 'head']):
                columns['description'] = col
            elif any(keyword in header for keyword in ['amount', 'total', '£', 'budget', 'cost']):
                if 'amount' not in columns:  # Take first amount column
                    columns['amount'] = col
        
        return columns
    
    def _extract_line_items(self, ws, start_row: int, columns: Dict[str, int]) -> List[Dict]:
        """Extract budget line items from rows"""
        line_items = []
        
        if 'description' not in columns or 'amount' not in columns:
            return line_items
        
        desc_col = columns['description']
        amount_col = columns['amount']
        
        for row_idx in range(start_row, min(start_row + 200, ws.max_row + 1)):
            desc = ws.cell(row=row_idx, column=desc_col).value
            amount_cell = ws.cell(row=row_idx, column=amount_col).value
            
            if not desc:
                continue
            
            desc_str = str(desc).strip()
            
            # Skip total rows, empty rows, headers
            if not desc_str or len(desc_str) < 2:
                continue
            if any(word in desc_str.lower() for word in ['total', 'grand total', 'sub-total', 'subtotal']):
                continue
            
            # Extract amount
            amount = self._parse_amount(amount_cell)
            if amount is None or amount == 0:
                continue
            
            # Categorize
            category = self._categorize_item(desc_str)
            
            line_items.append({
                'description': desc_str,
                'amount': amount,
                'category': category
            })
        
        return line_items
    
    def _parse_amount(self, value: Any) -> Optional[float]:
        """Parse amount from cell value"""
        if value is None:
            return None
        
        # If already a number
        if isinstance(value, (int, float)):
            return float(value)
        
        # Parse from string
        value_str = str(value).replace(',', '').replace('£', '').replace('$', '').strip()
        
        # Remove parentheses (negative numbers)
        is_negative = '(' in value_str
        value_str = value_str.replace('(', '').replace(')', '')
        
        try:
            amount = float(value_str)
            return -amount if is_negative else amount
        except:
            return None
    
    def _categorize_item(self, description: str) -> str:
        """Categorize budget line item"""
        desc_lower = description.lower()
        
        if any(word in desc_lower for word in ['reserve', 'sinking']):
            return 'reserve_fund'
        elif any(word in desc_lower for word in ['insurance', 'broker']):
            return 'insurance'
        elif any(word in desc_lower for word in ['management', 'agent', 'fee']):
            return 'management'
        elif any(word in desc_lower for word in ['cleaning', 'porter', 'caretaker']):
            return 'cleaning'
        elif any(word in desc_lower for word in ['lift', 'elevator']):
            return 'lift'
        elif any(word in desc_lower for word in ['garden', 'landscape']):
            return 'gardening'
        elif any(word in desc_lower for word in ['repair', 'maintenance']):
            return 'repairs'
        elif any(word in desc_lower for word in ['utility', 'electric', 'water', 'gas']):
            return 'utilities'
        else:
            return 'other'


def enhance_budgets_with_amounts(budgets: List[Dict], client_folder: str) -> List[Dict]:
    """
    Enhance budget records with actual amounts from Excel files
    
    Args:
        budgets: List of budget records with document references
        client_folder: Path to client folder
        
    Returns:
        Enhanced budget records with amounts
    """
    import os
    
    parser = BudgetParser()
    
    # Create a map of budget files
    budget_files = {}
    for root, dirs, files in os.walk(client_folder):
        for file in files:
            if file.lower().endswith(('.xlsx', '.xls')):
                if 'budget' in file.lower():
                    file_path = os.path.join(root, file)
                    budget_files[file.lower()] = file_path
    
    print(f"  📊 Found {len(budget_files)} budget Excel files")
    
    # Parse each budget file and enhance records
    enhanced_budgets = []
    parsed_data = {}
    
    for file_name, file_path in budget_files.items():
        print(f"     Parsing: {os.path.basename(file_path)}")
        data = parser.parse_budget_file(file_path)
        if data.get('line_items'):
            print(f"        ✅ Extracted {len(data['line_items'])} line items, Total: £{data.get('total_service_charge', 0):,.2f}")
            parsed_data[file_name] = data
    
    # Match parsed data to budget records
    for budget in budgets:
        source_file = budget.get('source_document', '').lower()
        
        # Try to find matching parsed data
        matched_data = None
        for file_name, data in parsed_data.items():
            if os.path.basename(source_file) in file_name or file_name in source_file:
                matched_data = data
                break
        
        if matched_data:
            budget['total_amount'] = matched_data.get('total_service_charge', 0)
            budget['year'] = matched_data.get('year')
            # Store line items separately
            budget['_line_items'] = matched_data.get('line_items', [])
        
        enhanced_budgets.append(budget)
    
    return enhanced_budgets

