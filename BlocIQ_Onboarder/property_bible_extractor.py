"""
Enhanced Property Bible & Property Form Extractor
Extracts comprehensive building metadata including systems, units, contractors
"""

import openpyxl
import re
from typing import Dict, Any, Optional, List
from datetime import datetime


class PropertyBibleExtractor:
    """Extract comprehensive building data from Property Bible and Property Form"""
    
    def __init__(self):
        pass
    
    def extract_from_excel(self, file_path: str) -> Dict[str, Any]:
        """
        Extract comprehensive building data from Excel files
        Handles both Property Bible and Property Form formats
        """
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            
            # Determine file type
            first_cell = str(ws['A1'].value or '').lower()
            
            if 'property setup' in first_cell or 'new property' in first_cell:
                return self._extract_property_form(ws)
            elif 'property bible' in str(ws.title).lower() or self._looks_like_bible(ws):
                return self._extract_property_bible(ws)
            else:
                # Try both methods
                data = self._extract_property_form(ws)
                bible_data = self._extract_property_bible(ws)
                # Merge, preferring property form data
                return {**bible_data, **data}
                
        except Exception as e:
            print(f"Error extracting from {file_path}: {e}")
            return {}
    
    def _looks_like_bible(self, ws) -> bool:
        """Check if worksheet looks like a Property Bible"""
        # Look for typical Property Bible sections
        for row in ws.iter_rows(max_row=30, values_only=True):
            row_text = ' '.join(str(cell) for cell in row if cell).lower()
            if any(term in row_text for term in ['freeholder', 'insurance broker', 'health & safety', 'employee']):
                return True
        return False
    
    def _extract_property_form(self, ws) -> Dict[str, Any]:
        """
        Extract from Property Setup Form
        Format: Label | Value
        """
        data = {}
        
        for row in ws.iter_rows(min_row=1, max_row=100, values_only=True):
            if not row or len(row) < 2:
                continue
            
            label = str(row[0] or '').strip()
            value = str(row[1] or '').strip() if len(row) > 1 else ''
            
            if not label or not value or value.lower() in ['none', 'n/a']:
                continue
            
            label_lower = label.lower()
            
            # Building metadata
            if 'number of units' in label_lower:
                data['number_of_units'] = self._extract_number(value)
            elif 'number of blocks' in label_lower:
                data['num_blocks'] = self._extract_number(value)
            elif 'building name' in label_lower or 'property name' in label_lower:
                data['building_name'] = value
            elif 'address' in label_lower and 'client' not in label_lower:
                data['address'] = value
            elif 'postcode' in label_lower:
                data['postcode'] = value
            
            # Financial
            elif 'management fee (ex vat)' in label_lower or 'management fee ex' in label_lower:
                data['management_fee_ex_vat'] = self._extract_currency(value)
            elif 'year end' in label_lower:
                data['year_end_date'] = value
            elif 'demand date' in label_lower:
                if 'demand_dates' not in data:
                    data['demand_dates'] = []
                data['demand_dates'].append(value)
            
            # Parties
            elif 'previous agent' in label_lower:
                data['previous_agents'] = value
            elif 'accountant' in label_lower:
                data['current_accountants'] = value
            elif 'insurance broker' in label_lower:
                data['insurance_broker'] = value
            
            # Ground Rent
            elif 'ground rent' in label_lower:
                data['ground_rent_applicable'] = 'yes' in value.lower() or '£' in value
                data['ground_rent_charges'] = value
        
        # Infer systems from document presence (will be enhanced by bible)
        data['_from_property_form'] = True
        
        return data
    
    def _extract_property_bible(self, ws) -> Dict[str, Any]:
        """
        Extract from Property Bible
        Format: Various sections with key information
        """
        data = {}
        
        # First pass: get building name from top
        building_name = str(ws['A1'].value or '').strip()
        if building_name and len(building_name) > 2:
            data['building_name'] = building_name
        
        current_section = None
        contractors = []
        staff = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=200, values_only=True), 1):
            if not row or not any(cell for cell in row):
                continue
            
            row_text = ' '.join(str(cell) for cell in row if cell).strip()
            row_lower = row_text.lower()
            
            # Identify sections
            if 'freeholder' in row_lower:
                current_section = 'freeholder'
            elif 'insurance' in row_lower and ('company' in row_lower or 'broker' in row_lower):
                current_section = 'insurance'
            elif 'health & safety' in row_lower or 'health and safety' in row_lower:
                current_section = 'compliance'
            elif 'employee' in row_lower:
                current_section = 'staff'
            
            # Extract based on section
            if 'insurance' in current_section if current_section else False:
                if '£' in row_text or 'premium' in row_lower:
                    # Extract insurance amount
                    amount = self._extract_currency(row_text)
                    if amount:
                        data['insurance_premium'] = amount
                
                # Extract renewal date
                date_match = re.search(r'(\d{1,2}(?:st|nd|rd|th)?\s+\w+\s+\d{4})', row_text)
                if date_match:
                    data['insurance_renewal_date'] = date_match.group(1)
            
            # Extract contractors from various sections
            if any(term in row_lower for term in ['lift', 'heating', 'cleaning', 'boiler', 'porter', 'gardening']):
                contractor_info = self._extract_contractor_from_row(row, row_lower)
                if contractor_info:
                    contractors.append(contractor_info)
            
            # Extract staff
            if current_section == 'staff':
                staff_info = self._extract_staff_from_row(row)
                if staff_info:
                    staff.append(staff_info)
        
        # Infer systems from contractors found
        data['has_lifts'] = any('lift' in c.get('service', '').lower() for c in contractors)
        data['has_communal_heating'] = any('heating' in c.get('service', '').lower() or 'boiler' in c.get('service', '').lower() for c in contractors)
        data['has_cleaning'] = any('clean' in c.get('service', '').lower() or 'porter' in c.get('service', '').lower() for c in contractors)
        data['has_gardening'] = any('garden' in c.get('service', '').lower() for c in contractors)
        
        if contractors:
            data['contractors'] = contractors
        if staff:
            data['staff'] = staff
        
        data['_from_property_bible'] = True
        
        return data
    
    def _extract_contractor_from_row(self, row, row_lower: str) -> Optional[Dict]:
        """Extract contractor information from a row"""
        contractor = {}
        
        # Determine service type
        if 'lift' in row_lower:
            contractor['service'] = 'lifts'
        elif 'heating' in row_lower or 'boiler' in row_lower:
            contractor['service'] = 'heating'
        elif 'clean' in row_lower or 'porter' in row_lower:
            contractor['service'] = 'cleaning'
        elif 'garden' in row_lower:
            contractor['service'] = 'gardening'
        elif 'electrical' in row_lower:
            contractor['service'] = 'electrical'
        else:
            return None
        
        # Try to extract company name (usually in first cell with capital letters)
        for cell in row:
            if cell and isinstance(cell, str):
                cell_str = str(cell).strip()
                # Look for company names (capitalized, more than 3 chars)
                if re.match(r'^[A-Z][A-Za-z\s&]+$', cell_str) and len(cell_str) > 3:
                    contractor['company_name'] = cell_str
                    break
        
        # Extract dates if present
        for cell in row:
            if cell:
                cell_str = str(cell)
                date_match = re.search(r'(\d{1,2}(?:st|nd|rd|th)?\s+\w+\s+\d{4})', cell_str)
                if date_match:
                    contractor['last_date'] = date_match.group(1)
        
        return contractor if 'service' in contractor else None
    
    def _extract_staff_from_row(self, row) -> Optional[Dict]:
        """Extract staff information from a row"""
        staff = {}
        
        for cell in row:
            if cell and isinstance(cell, str):
                cell_str = str(cell).strip()
                
                # Name (capitalized)
                if re.match(r'^[A-Z][a-z]+\s+[A-Z][a-z]+', cell_str):
                    staff['name'] = cell_str
                
                # Date (start date)
                elif isinstance(cell, datetime):
                    staff['start_date'] = cell.strftime('%Y-%m-%d')
        
        return staff if 'name' in staff else None
    
    def _extract_number(self, text: str) -> Optional[int]:
        """Extract first number from text"""
        match = re.search(r'\d+', str(text))
        return int(match.group()) if match else None
    
    def _extract_currency(self, text: str) -> Optional[float]:
        """Extract currency amount from text"""
        # Remove commas and extract number after £ or just a number
        text = str(text).replace(',', '')
        match = re.search(r'[£$]?\s*(\d+\.?\d*)', text)
        if match:
            try:
                return float(match.group(1))
            except:
                pass
        return None


def enhance_building_with_property_files(building_data: Dict, client_folder: str) -> Dict:
    """
    Enhance building data by extracting from Property Bible and Property Form
    
    Args:
        building_data: Existing building data dict
        client_folder: Path to client folder
    
    Returns:
        Enhanced building data dict
    """
    import os
    
    extractor = PropertyBibleExtractor()
    
    # Look for Property Bible and Property Form
    for root, dirs, files in os.walk(client_folder):
        for file in files:
            if file.lower().endswith(('.xlsx', '.xls')):
                file_lower = file.lower()
                file_path = os.path.join(root, file)
                
                if 'property bible' in file_lower or 'property form' in file_lower or 'property setup' in file_lower:
                    print(f"  📖 Found property metadata file: {file}")
                    extracted_data = extractor.extract_from_excel(file_path)
                    
                    # Merge into building_data, but don't overwrite existing values
                    for key, value in extracted_data.items():
                        if value and (key not in building_data or not building_data.get(key)):
                            building_data[key] = value
                            print(f"     ✅ Extracted {key}: {value}")
    
    return building_data

