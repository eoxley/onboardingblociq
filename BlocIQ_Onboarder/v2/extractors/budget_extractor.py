"""
Budget Extractor - Parse ALL line items with amounts
====================================================
Extracts complete budget data from Excel files
"""

import openpyxl
import re
from typing import Dict, List, Any, Optional
from datetime import datetime


class BudgetExtractor:
    """
    Extract complete budget data from Excel budget files
    Parses ALL line items with amounts, year, demand dates
    """
    
    def extract(self, filepath: str, document: Dict) -> Optional[Dict]:
        """
        Extract budget data from Excel file
        
        Returns:
            {
                'sc_year_start': '2024-04-01',
                'sc_year_end': '2025-03-31',
                'budget_year': 2025,
                'issue_date': '2024-03-15',
                'status': 'approved',  # draft/approved/final
                'total_budget': 92786.00,
                'line_items': [
                    {
                        'code': 'UTIL-ELEC',
                        'description': 'Utilities - Electricity',
                        'annual_amount': 6000.00,
                        'notes': '',
                        'category': 'utilities'
                    },
                    ...
                ]
            }
        """
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            
            budget_data = {
                'source_document': document.get('filename'),
                'document_id': document.get('document_id'),
                'sc_year_start': None,
                'sc_year_end': None,
                'budget_year': None,
                'issue_date': None,
                'status': 'draft',
                'total_budget': 0,
                'line_items': []
            }
            
            # Extract year from filename or content
            budget_data['budget_year'] = self._extract_year(filepath, ws)
            
            # Find budget table structure
            header_row = self._find_budget_header(ws)
            
            if not header_row:
                return None
            
            # Extract column positions
            columns = self._identify_columns(ws, header_row)
            
            # Extract all line items
            line_items = self._extract_line_items(ws, header_row + 1, columns)
            
            budget_data['line_items'] = line_items
            budget_data['total_budget'] = sum(item['annual_amount'] for item in line_items)
            
            # Try to find issue date
            budget_data['issue_date'] = self._find_issue_date(ws)
            
            # Determine status from content
            budget_data['status'] = self._determine_status(ws, filepath)
            
            # Calculate SC year based on budget year
            if budget_data['budget_year']:
                year = budget_data['budget_year']
                # Assume April-March year (common for UK service charge)
                budget_data['sc_year_start'] = f"{year-1}-04-01"
                budget_data['sc_year_end'] = f"{year}-03-31"
            
            return budget_data if line_items else None
        
        except Exception as e:
            print(f"      Budget extraction error: {str(e)[:100]}")
            return None
    
    def _extract_year(self, filepath: str, ws) -> Optional[int]:
        """Extract budget year from filename or content"""
        filename = str(filepath).lower()
        
        # Try filename patterns: "YE24", "YE2024", "2024-2025", "2024"
        patterns = [
            r'ye\s*(\d{4})',  # YE2024
            r'ye\s*(\d{2})',  # YE24
            r'(\d{4})\s*-\s*\d{4}',  # 2024-2025
            r'budget\s+(\d{4})',  # Budget 2024
        ]
        
        for pattern in patterns:
            match = re.search(pattern, filename)
            if match:
                year = match.group(1)
                if len(year) == 2:
                    year = '20' + year
                return int(year)
        
        # Try worksheet content (first 20 rows)
        for row in ws.iter_rows(max_row=20, values_only=True):
            row_text = ' '.join(str(cell) for cell in row if cell).lower()
            for pattern in patterns:
                match = re.search(pattern, row_text)
                if match:
                    year = match.group(1)
                    if len(year) == 2:
                        year = '20' + year
                    return int(year)
        
        return None
    
    def _find_budget_header(self, ws) -> Optional[int]:
        """Find row with budget table headers - ENHANCED for varied formats"""
        for row_idx in range(1, min(50, ws.max_row + 1)):
            row_text = ' '.join(
                str(ws.cell(row=row_idx, column=col).value or '').lower()
                for col in range(1, min(15, ws.max_column + 1))
            )
            
            # Look for budget year patterns
            has_budget_year = any(term in row_text for term in ['budget ye', 'ye 20', 'budget year', 'ye mar', 'ye 31'])
            
            # Look for traditional budget headers
            has_description = any(term in row_text for term in ['description', 'item', 'expense', 'head', 'category'])
            has_amount = any(term in row_text for term in ['amount', 'budget', 'annual', '£', 'cost', 'total'])
            
            # Accept if either format found
            if (has_description and has_amount) or has_budget_year:
                return row_idx
        
        return None
    
    def _identify_columns(self, ws, header_row: int) -> Dict[str, int]:
        """Identify which columns contain what data - ENHANCED for multi-column budgets"""
        columns = {}
        amount_columns = []  # Track all amount columns to pick the right one
        
        for col in range(1, min(20, ws.max_column + 1)):
            header = str(ws.cell(row=header_row, column=col).value or '').lower().strip()
            
            # Code/reference column
            if any(term in header for term in ['code', 'ref', 'account']):
                if 'code' not in columns:
                    columns['code'] = col
            
            # Amount columns - track all of them
            if any(term in header for term in ['amount', 'budget', 'annual', '£', 'cost', 'ye']):
                amount_columns.append((col, header))
            
            # Notes/comments column
            if any(term in header for term in ['note', 'comment', 'pm comment']):
                columns['notes'] = col
        
        # Description column - assume column 1 if not explicitly found
        # Most budgets have description in first column
        if 'description' not in columns:
            columns['description'] = 1
        
        # Pick the BEST amount column (prefer most recent year or "budget")
        if amount_columns:
            # Look for column with current year (2025, 2026) or "budget"
            best_col = amount_columns[0][0]  # Default to first
            
            for col, header in amount_columns:
                # Prefer columns with 2025, 2026, or "budget" in header
                if any(year in header for year in ['2025', '2026', 'budget ye 31']):
                    best_col = col
                    break
            
            columns['amount'] = best_col
        
        return columns
    
    def _extract_line_items(self, ws, start_row: int, columns: Dict[str, int]) -> List[Dict]:
        """Extract all budget line items"""
        line_items = []
        
        if 'description' not in columns or 'amount' not in columns:
            return line_items
        
        desc_col = columns['description']
        amount_col = columns['amount']
        code_col = columns.get('code')
        notes_col = columns.get('notes')
        
        for row_idx in range(start_row, min(start_row + 300, ws.max_row + 1)):
            desc = ws.cell(row=row_idx, column=desc_col).value
            amount_cell = ws.cell(row=row_idx, column=amount_col).value
            
            # Skip if no description
            if not desc:
                continue
            
            desc_str = str(desc).strip()
            
            # Skip total rows, empty rows, section headers
            if not desc_str or len(desc_str) < 2:
                continue
            
            # Skip total rows
            if any(word in desc_str.lower() for word in [
                'total', 'grand total', 'sub-total', 'subtotal', 
                'budget total', 'sum', '====', '----'
            ]):
                continue
            
            # Skip section headers (ALL CAPS headings with no amounts)
            if desc_str.isupper() and not amount_cell:
                continue
            
            # Extract amount
            amount = self._parse_amount(amount_cell)
            if amount is None:
                continue
            
            # Get optional fields
            code = str(ws.cell(row=row_idx, column=code_col).value or '') if code_col else ''
            notes = str(ws.cell(row=row_idx, column=notes_col).value or '') if notes_col else ''
            
            # Categorize line item
            category = self._categorize_expense(desc_str)
            
            line_items.append({
                'code': code.strip(),
                'description': desc_str,
                'annual_amount': amount,
                'notes': notes.strip(),
                'category': category,
                'row_number': row_idx
            })
        
        return line_items
    
    def _parse_amount(self, value: Any) -> Optional[float]:
        """Parse amount from cell value"""
        if value is None:
            return None
        
        # Already a number
        if isinstance(value, (int, float)):
            return float(value)
        
        # Parse from string
        value_str = str(value).replace(',', '').replace('£', '').replace('$', '').strip()
        
        # Handle negative numbers in parentheses
        is_negative = '(' in value_str
        value_str = value_str.replace('(', '').replace(')', '')
        
        try:
            amount = float(value_str)
            return -amount if is_negative else amount
        except:
            return None
    
    def _categorize_expense(self, description: str) -> str:
        """Categorize budget line item"""
        desc_lower = description.lower()
        
        if any(term in desc_lower for term in ['electric', 'power', 'lighting']):
            return 'utilities'
        elif any(term in desc_lower for term in ['gas', 'heating', 'boiler']):
            return 'utilities'
        elif any(term in desc_lower for term in ['water', 'drainage']):
            return 'utilities'
        elif any(term in desc_lower for term in ['clean', 'porter', 'caretaker']):
            return 'cleaning'
        elif any(term in desc_lower for term in ['garden', 'landscape', 'grounds']):
            return 'gardening'
        elif any(term in desc_lower for term in ['lift', 'elevator']):
            return 'lifts'
        elif any(term in desc_lower for term in ['repair', 'maintenance', 'decorat']):
            return 'repairs_maintenance'
        elif any(term in desc_lower for term in ['insurance', 'premium']):
            return 'insurance'
        elif any(term in desc_lower for term in ['management', 'agent', 'managing']):
            return 'management'
        elif any(term in desc_lower for term in ['accountan', 'audit']):
            return 'professional_fees'
        elif any(term in desc_lower for term in ['company sec', 'co sec']):
            return 'professional_fees'
        elif any(term in desc_lower for term in ['health', 'safety', 'compliance']):
            return 'compliance'
        elif any(term in desc_lower for term in ['reserve', 'sinking', 'contingency']):
            return 'reserve_fund'
        else:
            return 'other'
    
    def _find_issue_date(self, ws) -> Optional[str]:
        """Try to find budget issue/approval date"""
        # Look for dates in first 30 rows
        for row in ws.iter_rows(max_row=30, values_only=True):
            for cell in row:
                if isinstance(cell, datetime):
                    return cell.strftime('%Y-%m-%d')
                
                # Look for date strings
                if cell and isinstance(cell, str):
                    date_match = re.search(r'(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})', str(cell))
                    if date_match:
                        # TODO: Parse and format properly
                        return date_match.group(1)
        
        return None
    
    def _determine_status(self, ws, filepath: str) -> str:
        """Determine if budget is draft, approved, or final"""
        filename = str(filepath).lower()
        
        # Check filename
        if 'final' in filename or 'approved' in filename:
            return 'final'
        elif 'draft' in filename:
            return 'draft'
        
        # Check content
        for row in ws.iter_rows(max_row=30, values_only=True):
            row_text = ' '.join(str(cell) for cell in row if cell).lower()
            if 'approved' in row_text or 'final' in row_text:
                return 'approved'
            if 'draft' in row_text:
                return 'draft'
        
        return 'draft'  # Default to draft if unsure

