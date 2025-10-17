"""
Leaseholder Schedule Extractor
===============================
Intelligently detects and extracts from leaseholder list files
Recognizes patterns: reference numbers, names, addresses, phones, unit descriptions
"""

import openpyxl
import re
from typing import Dict, List, Optional


class LeaseholderScheduleExtractor:
    """
    Extract complete leaseholder data from schedule/list files
    
    Intelligently detects format by looking for:
    - Reference numbers (219-01-001 style)
    - Name column
    - Address column
    - Phone column
    - Unit/Flat descriptions
    """
    
    def is_leaseholder_file(self, filepath: str, ws) -> bool:
        """
        Intelligent detection: Is this a leaseholder information file?
        
        Looks for indicators:
        - Reference number column (219-01-001 format)
        - Name column
        - Address column
        - Unit/Flat column
        - Balance/Account columns (optional)
        """
        # Check first 5 rows for patterns
        for row_idx in range(1, min(6, ws.max_row + 1)):
            row_text = ' '.join(
                str(ws.cell(row=row_idx, column=col).value or '').lower()
                for col in range(1, min(15, ws.max_column + 1))
            )
            
            # Strong indicators this is a leaseholder file
            has_reference = any(term in row_text for term in ['reference', 'ref no', 'unit ref'])
            has_name = any(term in row_text for term in ['name', 'leaseholder', 'tenant', 'owner'])
            has_address = 'address' in row_text
            has_unit = any(term in row_text for term in ['unit', 'flat', 'property'])
            
            # If has 3+ indicators, it's likely a leaseholder file
            score = sum([has_reference, has_name, has_address, has_unit])
            
            if score >= 3:
                return True
            
            # Also check for reference number pattern in data (219-01-001 style)
            if re.search(r'\d{3}-\d{2}-\d{3}', row_text):
                return True
        
        return False
    
    def extract(self, filepath: str, document: Dict) -> List[Dict]:
        """
        Extract all leaseholder data from schedule file
        
        Returns list of:
        {
            'unit_number': '219-01-001',
            'leaseholder_name': 'Marmotte Holdings Limited',
            'correspondence_address': 'Flat 1, 32-34 Connaught Square...',
            'phone': '07768 803 607',
            'email': None,
            'unit_description': 'Flat 1 32-34 Connaught Square',
            'status': 'Current',
            'balance': 0.00
        }
        """
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            
            # Intelligent check: Is this a leaseholder file?
            if not self.is_leaseholder_file(filepath, ws):
                return []
            
            # Find header row
            header_row = self._find_header(ws)
            
            if not header_row:
                return []
            
            # Identify columns
            columns = self._identify_columns(ws, header_row)
            
            # Extract all leaseholders
            leaseholders = self._extract_leaseholders(ws, header_row + 1, columns)
            
            print(f"      Leaseholder schedule: {len(leaseholders)} leaseholders extracted")
            
            return leaseholders
        
        except Exception as e:
            print(f"      Leaseholder schedule error: {str(e)[:100]}")
            return []
    
    def _find_header(self, ws) -> Optional[int]:
        """Find header row"""
        for row_idx in range(1, min(10, ws.max_row + 1)):
            row_text = ' '.join(
                str(ws.cell(row=row_idx, column=col).value or '').lower()
                for col in range(1, min(15, ws.max_column + 1))
            )
            
            # Look for header indicators
            has_name = 'name' in row_text
            has_reference = 'reference' in row_text or 'ref' in row_text
            
            if has_name and has_reference:
                return row_idx
            
            # Also accept if has name + address + unit
            if 'name' in row_text and 'address' in row_text and 'unit' in row_text:
                return row_idx
        
        return None
    
    def _identify_columns(self, ws, header_row: int) -> Dict[str, int]:
        """Identify column positions"""
        columns = {}
        
        for col in range(1, min(20, ws.max_column + 1)):
            header = str(ws.cell(row=header_row, column=col).value or '').lower().strip()
            
            if 'reference' in header or header == 'ref':
                columns['reference'] = col
            
            elif header == 'name' or 'leaseholder' in header or 'tenant' in header:
                if 'name' not in columns:
                    columns['name'] = col
            
            elif 'address' in header:
                columns['address'] = col
            
            elif 'telephone' in header or 'phone' in header or 'mobile' in header:
                columns['phone'] = col
            
            elif 'email' in header or 'e-mail' in header:
                columns['email'] = col
            
            elif header == 'unit' or 'unit description' in header or 'flat' in header:
                if 'unit_desc' not in columns:
                    columns['unit_desc'] = col
            
            elif 'balance' in header or 'arrears' in header:
                columns['balance'] = col
            
            elif 'status' in header:
                if 'status' not in columns:
                    columns['status'] = col
        
        return columns
    
    def _extract_leaseholders(self, ws, start_row: int, columns: Dict[str, int]) -> List[Dict]:
        """Extract all leaseholder records"""
        leaseholders = []
        
        if 'reference' not in columns and 'name' not in columns:
            return []
        
        ref_col = columns.get('reference')
        name_col = columns.get('name')
        address_col = columns.get('address')
        phone_col = columns.get('phone')
        email_col = columns.get('email')
        unit_col = columns.get('unit_desc')
        status_col = columns.get('status')
        balance_col = columns.get('balance')
        
        for row_idx in range(start_row, min(start_row + 200, ws.max_row + 1)):
            # Get reference or name (must have one)
            reference = ws.cell(row=row_idx, column=ref_col).value if ref_col else None
            name = ws.cell(row=row_idx, column=name_col).value if name_col else None
            
            if not reference and not name:
                continue
            
            # Skip if empty row
            if not reference and not name:
                continue
            
            # Extract unit number from reference or description
            unit_number = None
            if reference:
                unit_number = str(reference).strip()
            elif unit_col:
                unit_desc = ws.cell(row=row_idx, column=unit_col).value
                if unit_desc:
                    # Extract from "Flat 1 32-34 Connaught Square"
                    flat_match = re.search(r'flat\s*(\d+)', str(unit_desc), re.IGNORECASE)
                    if flat_match:
                        unit_number = f"Flat {flat_match.group(1)}"
            
            if not unit_number:
                continue
            
            # Get all fields
            leaseholder = {
                'unit_number': unit_number,
                'leaseholder_name': str(name).strip() if name else None,
                'correspondence_address': str(ws.cell(row=row_idx, column=address_col).value or '').strip() if address_col else None,
                'phone': str(ws.cell(row=row_idx, column=phone_col).value or '').strip() if phone_col else None,
                'email': str(ws.cell(row=row_idx, column=email_col).value or '').strip() if email_col else None,
                'unit_description': str(ws.cell(row=row_idx, column=unit_col).value or '').strip() if unit_col else None,
                'status': str(ws.cell(row=row_idx, column=status_col).value or '').strip() if status_col else None,
                'balance': ws.cell(row=row_idx, column=balance_col).value if balance_col else None,
                'source': 'leaseholder_schedule',
                'authority_score': 1.0  # Leaseholder schedules are authoritative
            }
            
            # Clean up empty strings
            for key in ['correspondence_address', 'phone', 'email', 'unit_description']:
                if leaseholder.get(key) == '':
                    leaseholder[key] = None
            
            leaseholders.append(leaseholder)
        
        return leaseholders

