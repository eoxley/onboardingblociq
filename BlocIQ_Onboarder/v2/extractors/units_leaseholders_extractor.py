"""
Units & Leaseholders Unified Extractor
=======================================
Creates complete unit table with ALL fields:
- Unit number, floor
- Leaseholder name(s)
- Correspondence address
- Email, phone
- Apportionment %

Sources: Apportionment files, leases, property forms, correspondence
Conflict resolution: Authority ranking
"""

import openpyxl
import re
from typing import Dict, List, Optional, Tuple
from collections import defaultdict


class UnitsLeaseholdersExtractor:
    """
    Extract and unify units + leaseholders data from multiple sources
    Handles conflicts with authority ranking
    """
    
    def __init__(self):
        self.units = {}  # unit_number -> unit_data
        self.data_sources = defaultdict(list)  # Track data provenance
    
    def extract_from_apportionment(self, filepath: str, document: Dict) -> List[Dict]:
        """
        Extract from apportionment Excel file
        Primary source for unit numbers and apportionment %
        """
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            
            # Find header row
            header_row = self._find_apportionment_header(ws)
            if not header_row:
                return []
            
            # Identify columns
            columns = self._identify_apportionment_columns(ws, header_row)
            
            # Extract units
            units_extracted = []
            
            for row_idx in range(header_row + 1, min(header_row + 200, ws.max_row + 1)):
                unit_num = ws.cell(row=row_idx, column=columns.get('unit', 1)).value
                
                if not unit_num or str(unit_num).strip() == '':
                    continue
                
                unit_num_str = str(unit_num).strip()
                
                # Skip total rows
                if any(word in unit_num_str.lower() for word in ['total', 'sum', '==']):
                    continue
                
                # Extract data
                unit_data = {
                    'unit_number': unit_num_str,
                    'floor': self._extract_floor_from_unit(unit_num_str),
                    'leaseholder_name': None,
                    'correspondence_address': None,
                    'email': None,
                    'phone': None,
                    'apportionment': None,
                    'source': 'apportionment',
                    'authority_score': 1.0  # Apportionment is authoritative
                }
                
                # Get leaseholder name if column exists
                if 'leaseholder' in columns:
                    name = ws.cell(row=row_idx, column=columns['leaseholder']).value
                    if name:
                        unit_data['leaseholder_name'] = str(name).strip()
                
                # Get apportionment percentage
                if 'apportionment' in columns:
                    apport = ws.cell(row=row_idx, column=columns['apportionment']).value
                    if apport:
                        # Parse percentage
                        unit_data['apportionment'] = self._parse_percentage(apport)
                
                # Get address if exists
                if 'address' in columns:
                    addr = ws.cell(row=row_idx, column=columns['address']).value
                    if addr:
                        unit_data['correspondence_address'] = str(addr).strip()
                
                units_extracted.append(unit_data)
                self._merge_unit_data(unit_data)
            
            print(f"      Apportionment: {len(units_extracted)} units extracted")
            return units_extracted
        
        except Exception as e:
            print(f"      Apportionment extraction error: {str(e)[:100]}")
            return []
    
    def _find_apportionment_header(self, ws) -> Optional[int]:
        """Find header row in apportionment file - ENHANCED for various formats"""
        for row_idx in range(1, min(30, ws.max_row + 1)):
            row_text = ' '.join(
                str(ws.cell(row=row_idx, column=col).value or '').lower()
                for col in range(1, min(15, ws.max_column + 1))
            )
            
            # Look for typical headers
            has_unit = any(term in row_text for term in ['unit', 'flat', 'property'])
            has_percentage = any(term in row_text for term in ['%', 'percent', 'apport', 'share', 'rate'])  # Added 'rate'!
            
            if has_unit and has_percentage:
                return row_idx
        
        return None
    
    def _identify_apportionment_columns(self, ws, header_row: int) -> Dict[str, int]:
        """Identify column positions - ENHANCED for various formats"""
        columns = {}
        
        for col in range(1, min(20, ws.max_column + 1)):
            header = str(ws.cell(row=header_row, column=col).value or '').lower().strip()
            
            # Unit reference/number column
            if any(term in header for term in ['unit ref', 'unit', 'flat', 'property']) and 'description' not in header:
                if 'unit' not in columns:
                    columns['unit'] = col
            
            # Unit description (may contain flat number too)
            elif 'description' in header:
                columns['description'] = col
            
            # Leaseholder name
            elif any(term in header for term in ['name', 'leaseholder', 'owner', 'tenant']):
                columns['leaseholder'] = col
            
            # Apportionment percentage - ADDED 'rate'!
            elif any(term in header for term in ['%', 'percent', 'apport', 'share', 'rate']):
                if 'apportionment' not in columns:  # Take first percentage column
                    columns['apportionment'] = col
            
            # Address
            elif any(term in header for term in ['address', 'correspondence']):
                columns['address'] = col
            
            # Email
            elif any(term in header for term in ['email', 'e-mail']):
                columns['email'] = col
            
            # Phone
            elif any(term in header for term in ['phone', 'tel', 'mobile']):
                columns['phone'] = col
        
        return columns
    
    def _parse_percentage(self, value) -> Optional[float]:
        """Parse apportionment percentage"""
        if value is None:
            return None
        
        # Already a number
        if isinstance(value, (int, float)):
            val = float(value)
            # If > 1, assume it's already a percentage (e.g., 5.5 for 5.5%)
            # If <= 1, assume it's a decimal (e.g., 0.055 for 5.5%)
            if val <= 1:
                return val * 100
            return val
        
        # Parse from string
        value_str = str(value).strip().replace('%', '')
        try:
            val = float(value_str)
            if val <= 1:
                return val * 100
            return val
        except:
            return None
    
    def _extract_floor_from_unit(self, unit_number: str) -> Optional[int]:
        """
        Extract floor number from unit number
        e.g., "Flat 1" -> Ground floor (0)
        "Flat 15" -> First floor (1)
        "Flat 25" -> Second floor (2)
        """
        # Try to find explicit floor mentions
        floor_patterns = [
            (r'\b(?:ground|gf|g)\b', 0),
            (r'\b(?:first|1st|1f)\b', 1),
            (r'\b(?:second|2nd|2f)\b', 2),
            (r'\b(?:third|3rd|3f)\b', 3),
            (r'\b(?:fourth|4th|4f)\b', 4),
        ]
        
        unit_lower = unit_number.lower()
        for pattern, floor_num in floor_patterns:
            if re.search(pattern, unit_lower):
                return floor_num
        
        # Infer from unit number (e.g., Flat 1-9 = ground, 10-19 = first)
        unit_match = re.search(r'(\d+)', unit_number)
        if unit_match:
            unit_num = int(unit_match.group(1))
            return unit_num // 10  # Simple heuristic
        
        return None
    
    def _merge_unit_data(self, new_data: Dict):
        """
        Merge unit data with conflict resolution
        Higher authority score wins
        """
        unit_num = new_data['unit_number']
        
        if unit_num not in self.units:
            self.units[unit_num] = new_data
            return
        
        existing = self.units[unit_num]
        
        # Merge fields based on authority score
        for field in ['leaseholder_name', 'correspondence_address', 'email', 'phone', 'apportionment']:
            new_value = new_data.get(field)
            existing_value = existing.get(field)
            
            # If new value exists and (existing doesn't or new has higher authority)
            if new_value:
                if not existing_value or new_data['authority_score'] >= existing['authority_score']:
                    existing[field] = new_value
    
    def get_all_units(self) -> List[Dict]:
        """Get final unified units list"""
        units_list = list(self.units.values())
        
        # Sort by unit number
        def sort_key(unit):
            unit_num = unit['unit_number']
            # Extract numeric part
            match = re.search(r'(\d+)', unit_num)
            if match:
                return int(match.group(1))
            return 0
        
        units_list.sort(key=sort_key)
        
        return units_list
    
    def extract_from_lease(self, lease_text: str, title_number: str) -> Optional[Dict]:
        """
        Extract unit/leaseholder from lease document
        Lower authority than apportionment
        """
        # Extract leaseholder name (Tenant)
        leaseholder = self._extract_tenant_name(lease_text)
        
        # Extract property description
        unit_desc = self._extract_property_description(lease_text)
        
        if not leaseholder and not unit_desc:
            return None
        
        return {
            'unit_number': unit_desc or f'Title {title_number}',
            'leaseholder_name': leaseholder,
            'apportionment': None,  # Extract from schedule if available
            'source': 'lease',
            'authority_score': 0.7
        }
    
    def _extract_tenant_name(self, text: str) -> Optional[str]:
        """Extract tenant/leaseholder name from lease"""
        patterns = [
            r'tenant:?\s*([A-Z][A-Za-z\s]+(?:and|&)[A-Za-z\s]+)',  # Multiple names
            r'lessee:?\s*([A-Z][A-Za-z\s]+)',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text[:3000], re.IGNORECASE)
            if match:
                return match.group(1).strip()[:100]
        
        return None
    
    def _extract_property_description(self, text: str) -> Optional[str]:
        """Extract property/unit description"""
        patterns = [
            r'demised\s+premises:?\s*([A-Za-z0-9\s,]+)',
            r'flat\s+(\d+[A-Z]?)',
            r'apartment\s+(\d+[A-Z]?)',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text[:3000], re.IGNORECASE)
            if match:
                return match.group(1).strip()[:100]
        
        return None
    
    def print_summary(self):
        """Print extraction summary"""
        units = self.get_all_units()
        
        print(f"\n🏠 UNITS & LEASEHOLDERS SUMMARY:")
        print(f"   Total units: {len(units)}")
        
        with_leaseholders = len([u for u in units if u.get('leaseholder_name')])
        with_apportionment = len([u for u in units if u.get('apportionment')])
        
        print(f"   With leaseholder names: {with_leaseholders}")
        print(f"   With apportionment: {with_apportionment}")
        print()
        
        # Show sample
        for unit in units[:10]:
            print(f"   • {unit['unit_number']}: {unit.get('leaseholder_name', 'Unknown')} - {unit.get('apportionment', '?')}%")
        
        if len(units) > 10:
            print(f"   ... and {len(units) - 10} more units")

